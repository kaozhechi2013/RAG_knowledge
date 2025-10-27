"""
Excel/CSV 表格转自然语言文本工具

用途：
- 将研发计划、人员任务等结构化表格转换为适合知识库检索的自然语言描述
- 默认按行生成段落，每一列以 "列名：值" 的形式输出，可选自定义选项

示例：
python table_to_text.py "数字工程所2025年研发计划.xlsx" --sheet "Sheet1"

输出：
- 默认会在同级目录生成同名 txt 文件，例如 `数字工程所2025年研发计划_clean.txt`
- 每一行输出一个段落，列名和值之间使用全角冒号，列之间使用换行，段落之间空一行

依赖：
- Excel 文件需要安装 `openpyxl` (pip install openpyxl)
- CSV 文件使用内置 csv 模块，无需额外依赖
"""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Sequence

try:  # 尝试按需导入 openpyxl（Excel 支持）
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover - 在未安装 openpyxl 的环境下提醒
    load_workbook = None  # type: ignore


@dataclass
class TableRow:
    """结构化行数据。"""

    values: Sequence[str]

    def to_paragraph(self, headers: Sequence[str], bullet: bool = True) -> str:
        """将当前行渲染为自然语言段落。"""

        fragments: List[str] = []
        for header, raw in zip(headers, self.values):
            header = header.strip()
            value = raw.strip()
            if not header or not value:
                continue

            prefix = "- " if bullet else ""
            fragments.append(f"{prefix}{header}：{value}")

        return "\n".join(fragments)


def read_excel(path: Path, sheet_name: str | None = None) -> tuple[List[str], List[TableRow]]:
    if load_workbook is None:
        raise RuntimeError(
            "未检测到 openpyxl，请先运行 `pip install openpyxl` 再处理 Excel 文件。"
        )

    workbook = load_workbook(path, read_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
    except KeyError as exc:  # pragma: no cover - 只有 sheet 名错时触发
        workbook.close()
        raise RuntimeError(f"工作表不存在: {sheet_name}") from exc

    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        raise RuntimeError("表格为空，没有可转换的内容。")

    headers = [normalize_cell(cell) for cell in rows[0]]
    data_rows = [TableRow([normalize_cell(cell) for cell in row]) for row in rows[1:]]
    return headers, data_rows


def read_csv(path: Path, encoding: str = "utf-8-sig") -> tuple[List[str], List[TableRow]]:
    with path.open("r", encoding=encoding, newline="") as f:
        reader = csv.reader(f)
        try:
            headers = [normalize_cell(cell) for cell in next(reader)]
        except StopIteration:
            raise RuntimeError("CSV 文件为空，没有可转换的内容。")

        data_rows = [TableRow([normalize_cell(cell) for cell in row]) for row in reader]
    return headers, data_rows


def normalize_cell(cell: object | None) -> str:
    if cell is None:
        return ""
    if isinstance(cell, float):
        # 如果是整数值的浮点数，去掉小数点
        if cell.is_integer():
            return str(int(cell))
        return f"{cell:.6g}"  # 保留重要数字
    return str(cell).strip()


def build_output_path(input_path: Path, output: str | None) -> Path:
    if output:
        output_path = Path(output)
        if output_path.is_dir():
            return output_path / f"{input_path.stem}_clean.txt"
        return output_path

    return input_path.with_name(f"{input_path.stem}_clean.txt")


def render_rows(headers: Sequence[str], rows: Iterable[TableRow], bullet: bool) -> str:
    paragraphs: List[str] = []
    for row in rows:
        paragraph = row.to_paragraph(headers, bullet=bullet)
        if paragraph:
            paragraphs.append(paragraph)

    return "\n\n".join(paragraphs)


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="将 Excel/CSV 表格转换为自然语言文本")
    parser.add_argument("input", help="要处理的 Excel/CSV 文件路径")
    parser.add_argument("--sheet", help="Excel 工作表名称，默认取活动工作表")
    parser.add_argument(
        "--output",
        help="输出文件路径或目录（默认与输入文件同目录，同名 *_clean.txt）",
    )
    parser.add_argument(
        "--no-bullet",
        action="store_true",
        help="输出时不使用前缀短横线，改为直接“列名：值”",
    )
    parser.add_argument(
        "--encoding",
        default="utf-8-sig",
        help="读取 CSV 时的编码，默认 utf-8-sig",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    input_path = Path(args.input).expanduser().resolve()

    if not input_path.exists():
        print(f"❌ 输入文件不存在: {input_path}")
        return 1

    try:
        if input_path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            headers, rows = read_excel(input_path, sheet_name=args.sheet)
        elif input_path.suffix.lower() in {".csv"}:
            headers, rows = read_csv(input_path, encoding=args.encoding)
        else:
            print("❌ 仅支持 Excel (.xlsx, .xlsm) 或 CSV 文件。")
            return 1
    except Exception as exc:
        print(f"❌ 读取表格失败: {exc}")
        return 1

    paragraphs = render_rows(headers, rows, bullet=not args.no_bullet)
    if not paragraphs:
        print("⚠️ 表格中没有有效内容。")
        return 0

    output_path = build_output_path(input_path, args.output)
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(paragraphs, encoding="utf-8")
    except Exception as exc:  # pragma: no cover - IO 错误时提示
        print(f"❌ 写入文件失败: {exc}")
        return 1

    print(f"✅ 已生成自然语言文件: {output_path}")
    paragraph_count = paragraphs.count("\n\n") + 1 if paragraphs else 0
    print(f"   段落数: {paragraph_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
